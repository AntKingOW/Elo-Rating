"""
OWCS 2024-2026 Combined Elo — 2026 Active Teams Only

Builds a combined 2024-2026 phase chart using only teams that are active in
2026 Stage 1. Team continuity is resolved from `OWCS Team History.xlsx`:
names appearing in the same row are treated as the same team, and the display
name follows the 2026 S1 name for that row when available.

Usage:
    python visualize_combined_elo_2026.py
"""

from __future__ import annotations

import colorsys
import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
HISTORY_2024 = ROOT / "OWCS_2024_GLOBAL_ELO_HISTORY.csv"
HISTORY_2025 = ROOT / "OWCS_2025_GLOBAL_ELO_HISTORY.csv"
HISTORY_2026 = ROOT / "OWCS_2026_GLOBAL_ELO_HISTORY.csv"
FINAL_2026 = ROOT / "OWCS_2026_GLOBAL_ELO_FINAL.csv"
TEAM_HISTORY_XLSX = ROOT / "OWCS Team History.xlsx"
OUTPUT_HTML = ROOT / "OWCS_COMBINED_ELO_2024_2026.html"

REGION_ORDER = ["Korea", "Japan", "Pacific", "China", "EMEA", "NA"]
DEFAULT_VISIBLE_TEAMS = 30

PHASE_DEFS: list[tuple[str, list[str]]] = [
    ("2024 S1+S2", [
        "owcs_2024_asia_s1_pacific",
        "owcs_2024_na_s1_groups",
        "owcs_2024_emea_s1_groups",
        "owcs_2024_na_s1_main",
        "owcs_2024_emea_s1_main",
        "owcs_2024_asia_s1_japan",
        "owcs_2024_asia_s1_korea",
        "owcs_2024_asia_s1_wildcard",
        "owcs_2024_asia_s1_main",
        "owcs_2024_na_s2_groups",
        "owcs_2024_emea_s2_groups",
        "owcs_2024_na_s2_main",
        "owcs_2024_emea_s2_main",
    ]),
    ("Dallas\nMajor", ["owcs_2024_dallas_major"]),
    ("EWC 2024", ["ewc_2024"]),
    ("2024 S3–S4", [
        "owcs_2024_asia_s2_pacific",
        "owcs_2024_na_s3_groups",
        "owcs_2024_emea_s3_groups",
        "owcs_2024_na_s3_main",
        "owcs_2024_emea_s3_main",
        "owcs_2024_asia_s2_korea",
        "owcs_2024_asia_s2_japan",
        "owcs_2024_asia_s2_wildcard",
        "owcs_2024_na_s4_groups",
        "owcs_2024_emea_s4_groups",
        "owcs_2024_asia_s2_main",
        "owcs_2024_na_s4_main",
        "owcs_2024_emea_s4_main",
        "faceit_2024_s3_na_master",
        "faceit_2024_s3_emea_master",
    ]),
    ("2024\nWorld Finals", ["owcs_2024_world_finals"]),
    ("2025 Stage 1", [
        "owcs_2025_asia_s1_japan",
        "owcs_2025_asia_s1_pacific",
        "owcs_2025_asia_s1_korea",
        "owcs_2025_na_s1",
        "owcs_2025_emea_s1",
        "owcs_2025_asia_s1_main",
        "owcs_2025_china_s1",
    ]),
    ("Champions\nClash", ["owcs_2025_champions_clash"]),
    ("2025 Stage 2", [
        "owcs_2025_asia_s2_korea",
        "owcs_2025_na_s2",
        "owcs_2025_emea_s2",
        "owcs_2025_asia_s2_japan",
        "owcs_2025_asia_s2_pacific",
        "owcs_2025_china_s2",
    ]),
    ("Midseason\nChampionship", ["owcs_2025_midseason"]),
    ("2025 Stage 3", [
        "owcs_2025_asia_s3_japan",
        "owcs_2025_asia_s3_pacific",
        "owcs_2025_asia_s3_korea",
        "owcs_2025_na_s3",
        "owcs_2025_emea_s3",
        "owcs_2025_china_s3",
        "owcs_2025_apac_championship",
        "owcs_2025_korea_road_to_wf",
    ]),
    ("2025\nWorld Finals", ["owcs_2025_world_finals"]),
    ("2026 Stage 1", [
        "owcs_2026_asia_s1_japan",
        "owcs_2026_asia_s1_korea",
        "owcs_2026_asia_s1_pacific",
        "owcs_2026_china_s1",
        "owcs_2026_emea_s1",
        "owcs_2026_na_s1",
    ]),
]

BOUNDARY_AFTER_PHASES = [4, 10]

# Team History workbook uses one row per team lineage. These are the columns whose
# values should be treated as the 2026 S1 display name for the sheet.
S1_2026_COLUMN = {
    "Korea": 14,
    "Japan": 14,
    "Pacific": 14,  # workbook year header is mistyped, but col 14 is the 2026 S1 block
    "China": 10,
    "EMEA": 18,
    "NA": 18,
}

DISPLAY_FIXES = {
    "Crazy Racoon": "Crazy Raccoon",
    "Varrel": "VARREL",
    "ENTER FORCE.36": "Enter Force.36",
    "Tokyo Ta1yos": "Tokyo Ta1yo's",
    "Naïve Piggy": "Naive Piggy",
    "Twisted Mind": "Twisted Minds",
    "Anyone's Legend": "Anyone's Legend",
    "ONSIDE Gaming": "ONSIDE Gaming",
    "JDG Gaming": "JDG Gaming",
    "Fury": "FURY",
    "Team Secret": "Team Secret",
}

DATA_ALIAS_FIXES = {
    "Crazy Raccoon": "Crazy Raccoon",
    "VARREL": "VARREL",
    "Enter Force.36": "Enter Force.36",
    "Tokyo Ta1yo's": "Tokyo Ta1yo's",
    "Naive Piggy": "Naive Piggy",
    "Twisted Minds": "Twisted Minds",
    "anyone's legend": "Anyone's Legend",
    "ONSIDE GAMING": "ONSIDE Gaming",
    "JD Gaming": "JDG Gaming",
    "FURY": "FURY",
    "disguised": "Disguised",
}

FIXED_COLORS = {
    "Crazy Raccoon": "#E30613",
    "Team Falcons": "#00A651",
    "ZETA DIVISION": "#000000",
    "T1": "#E2012D",
    "VARREL": "#00B1EB",
    "Please Not Hero Ban": "#A8E6CF",
    "Weibo Gaming": "#FF4500",
    "JDG Gaming": "#C8102E",
    "Twisted Minds": "#FA4B68",
    "Team Peps": "#FEF100",
    "Virtus.pro": "#F65101",
    "Al Qadsiah": "#FDE100",
    "Team Liquid": "#00274D",
    "Spacestation Gaming": "#FDB827",
    "Dallas Fuel": "#0F5BA7",
}


def _hsl_to_hex(h: float, s: float, l: float) -> str:
    r, g, b = colorsys.hls_to_rgb(h / 360, l, s)
    return f"#{int(r * 255):02x}{int(g * 255):02x}{int(b * 255):02x}"


def auto_color(name: str) -> str:
    if name in FIXED_COLORS:
        return FIXED_COLORS[name]
    seed = sum((i + 1) * ord(ch) for i, ch in enumerate(name))
    hue = seed % 360
    sat = 0.65 + ((seed // 17) % 15) / 100
    light = 0.44 + ((seed // 29) % 18) / 100
    return _hsl_to_hex(hue, min(sat, 0.82), min(light, 0.62))


def clean_cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith(">"):
        return None
    if text in {"Team Name", "S1", "S2", "S3", "S4"}:
        return None
    return DISPLAY_FIXES.get(text, text)


def canonicalize_data_name(name: str) -> str:
    return DATA_ALIAS_FIXES.get(name, DISPLAY_FIXES.get(name, name))


def load_team_history_groups() -> tuple[dict[str, str], dict[str, str], set[str]]:
    wb = load_workbook(TEAM_HISTORY_XLSX, data_only=True)

    alias_to_canonical: dict[str, str] = {}
    canonical_to_region: dict[str, str] = {}
    canonicals: set[str] = set()

    for sheet_name in REGION_ORDER:
        ws = wb[sheet_name]
        canonical_col = S1_2026_COLUMN[sheet_name]

        for row_idx in range(5, ws.max_row + 1):
            canonical_raw = clean_cell(ws.cell(row_idx, canonical_col).value)
            if canonical_raw is None:
                continue

            canonical = canonical_raw
            canonicals.add(canonical)
            canonical_to_region[canonical] = sheet_name

            aliases: set[str] = {canonical}
            for col_idx in range(4, ws.max_column + 1):
                alias = clean_cell(ws.cell(row_idx, col_idx).value)
                if alias is not None:
                    aliases.add(alias)

            for alias in aliases:
                alias_to_canonical[alias] = canonical

    # Add standalone 2026 teams not covered by workbook rows.
    with FINAL_2026.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            actual = canonicalize_data_name(row["team"])
            if actual not in canonicals:
                canonicals.add(actual)
                canonical_to_region[actual] = row["home_region"]
            alias_to_canonical.setdefault(actual, actual)
            alias_to_canonical.setdefault(row["team"], actual)

    return alias_to_canonical, canonical_to_region, canonicals


def load_data() -> list[dict]:
    event_to_phase: dict[str, int] = {}
    for phase_idx, (_, event_ids) in enumerate(PHASE_DEFS):
        for event_id in event_ids:
            event_to_phase[event_id] = phase_idx

    alias_to_canonical, canonical_to_region, active_canonicals = load_team_history_groups()
    team_phase_elo: dict[str, dict[int, float]] = defaultdict(dict)
    num_phases = len(PHASE_DEFS)

    def process_csv(path: Path) -> None:
        if not path.exists():
            print(f"  [WARN] {path.name} not found - skipping")
            return

        for row in csv.DictReader(path.open(encoding="utf-8")):
            raw_team = row["team"]
            team = alias_to_canonical.get(raw_team)
            if team is None:
                team = alias_to_canonical.get(canonicalize_data_name(raw_team))
            if team is None or team not in active_canonicals:
                continue

            phase = event_to_phase.get(row["event_id"])
            if phase is None:
                continue
            team_phase_elo[team][phase] = float(row["elo_after"])

    process_csv(HISTORY_2024)
    process_csv(HISTORY_2025)
    process_csv(HISTORY_2026)

    results: list[dict] = []
    for team in sorted(active_canonicals):
        elo_map = team_phase_elo.get(team, {})
        phase_elos: list[float | None] = []
        current: float | None = None
        first_seen = False

        for idx in range(num_phases):
            if idx in elo_map:
                current = round(elo_map[idx], 1)
                first_seen = True
            elif not first_seen:
                current = None
            phase_elos.append(current)

        final_elo = next((v for v in reversed(phase_elos) if v is not None), None)
        if final_elo is None:
            continue

        results.append({
            "name": team,
            "region": canonical_to_region.get(team, "Intl"),
            "color": auto_color(team),
            "phaseElo": phase_elos,
            "playedPhases": sorted(elo_map.keys()),
            "finalElo": final_elo,
        })

    results.sort(key=lambda item: item["finalElo"], reverse=True)
    for idx, team in enumerate(results):
        team["defaultHidden"] = idx >= DEFAULT_VISIBLE_TEAMS
    return results


def build_html(teams: list[dict]) -> str:
    phase_labels = [label for label, _ in PHASE_DEFS]
    teams_json = json.dumps(teams, ensure_ascii=False)
    phase_labels_json = json.dumps(phase_labels, ensure_ascii=False)
    boundaries_json = json.dumps(BOUNDARY_AFTER_PHASES, ensure_ascii=False)
    regions_json = json.dumps(REGION_ORDER, ensure_ascii=False)

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>OWCS 2024-2026 Combined Elo</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  *, *::before, *::after {{ box-sizing: border-box; }}
  :root {{
    --bg: #ffffff;
    --surface: #ffffff;
    --border: rgba(0, 0, 0, 0.12);
    --text: #1a1a1a;
    --muted: #555555;
    --accent: #2563eb;
    --shadow: 0 4px 20px rgba(0,0,0,0.08);
  }}
  body {{
    margin: 0;
    background: #f5f7fa;
    color: var(--text);
    font-family: "Segoe UI", "Noto Sans KR", sans-serif;
    min-height: 100vh;
  }}
  .shell {{
    max-width: 1400px;
    margin: 0 auto;
    padding: 32px 24px;
  }}
  header {{
    padding: 28px 32px;
    border: 1px solid var(--border);
    border-radius: 20px;
    background: #ffffff;
    box-shadow: var(--shadow);
    margin-bottom: 20px;
  }}
  .eyebrow {{
    color: #2563eb;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    margin-bottom: 8px;
  }}
  h1 {{
    margin: 0 0 8px;
    font-size: clamp(22px, 4vw, 40px);
    line-height: 1.1;
    color: #111111;
  }}
  .subtitle {{
    margin: 0;
    color: var(--muted);
    font-size: 14px;
  }}
  .card {{
    border: 1px solid var(--border);
    border-radius: 20px;
    background: #ffffff;
    box-shadow: var(--shadow);
    padding: 24px;
    margin-bottom: 20px;
  }}
  .card h2 {{
    margin: 0 0 4px;
    font-size: 16px;
    font-weight: 700;
    color: #111111;
  }}
  .card .desc {{
    margin: 0 0 18px;
    color: var(--muted);
    font-size: 13px;
  }}
  .chart-wrap {{
    position: relative;
    height: 620px;
  }}
  .region-filters {{
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    margin-top: 18px;
    margin-bottom: 18px;
  }}
  .region-filter {{
    border: 1px solid var(--border);
    border-radius: 999px;
    background: #fafafa;
    color: #333333;
    padding: 8px 14px;
    font-size: 12px;
    font-weight: 700;
    cursor: pointer;
    transition: background 0.15s, color 0.15s, border-color 0.15s;
  }}
  .region-filter.active {{
    background: rgba(37,99,235,0.08);
    color: #2563eb;
    border-color: rgba(37,99,235,0.28);
  }}
  .legend-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(220px, 1fr));
    gap: 10px;
    margin-top: 22px;
  }}
  .legend-item {{
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 10px 14px;
    border: 1px solid var(--border);
    border-radius: 12px;
    background: #fafafa;
    font-size: 13px;
    cursor: pointer;
    transition: background 0.15s;
  }}
  .legend-item:hover {{
    background: #f0f0f0;
  }}
  .legend-item.hidden {{
    opacity: 0.35;
  }}
  .dot {{
    width: 12px;
    height: 12px;
    border-radius: 50%;
    flex-shrink: 0;
    border: 1px solid rgba(0,0,0,0.15);
  }}
  .legend-name {{
    flex: 1;
    font-weight: 600;
    color: #1a1a1a;
  }}
  .legend-elo {{
    color: var(--muted);
    font-size: 12px;
    font-weight: 700;
  }}
  .note {{
    margin-top: 18px;
    padding: 14px 18px;
    border: 1px solid rgba(37,99,235,0.2);
    border-radius: 12px;
    background: rgba(37,99,235,0.04);
    color: var(--muted);
    font-size: 12.5px;
    line-height: 1.7;
  }}
</style>
</head>
<body>
<div class="shell">

<header>
  <div class="eyebrow">OWCS Elo Rating — 통합 시즌 차트</div>
  <h1>2024 + 2025 + 2026 Combined Elo Trajectory</h1>
  <p class="subtitle">
    2026 Stage 1에 실제 출전 중인 팀만 표시 &mdash;
    팀 히스토리 엑셀 기준 같은 열은 하나의 팀으로 통합 &middot;
    수직 점선: 2024 / 2025 / 2026 시즌 경계
  </p>
</header>

<div class="card">
  <h2>Phase Elo Trend</h2>
  <p class="desc">2026 출전 팀만 대상입니다. 지역 선택 후 범례 클릭으로 팀 ON/OFF</p>
  <div class="chart-wrap">
    <canvas id="chart"></canvas>
  </div>
  <div class="region-filters" id="regionFilters"></div>
  <div class="legend-grid" id="legend"></div>
  <div class="note">
    <strong>표기 안내:</strong>
    `OWCS Team History.xlsx`의 같은 행은 동일 팀으로 취급했고, 표시 이름은 2026 S1 이름을 기준으로 맞췄습니다. 2026 S1에 출전한 팀만 차트에 남겼습니다.
  </div>
</div>

</div>

<script>
const PHASES = {phase_labels_json};
const TEAMS = {teams_json};
const BOUNDARIES = {boundaries_json};
const REGIONS = {regions_json};

const datasets = TEAMS.map(t => ({{
  label: t.name,
  data: t.phaseElo,
  borderColor: t.color,
  backgroundColor: t.color + "28",
  borderWidth: 2.5,
  pointRadius: PHASES.map((_, idx) => t.playedPhases.includes(idx) ? 5 : 0),
  pointHoverRadius: PHASES.map((_, idx) => t.playedPhases.includes(idx) ? 8 : 0),
  pointBackgroundColor: t.color,
  tension: 0,
  spanGaps: false,
  hidden: !!t.defaultHidden,
}}));

const manualHidden = TEAMS.map(t => !!t.defaultHidden);
let activeRegions = new Set(REGIONS);

const boundaryPlugin = {{
  id: "yearBoundary",
  afterDraw(chart) {{
    const {{ ctx, chartArea: {{ top, bottom, left, right }}, scales: {{ x }} }} = chart;
    const linePositions = BOUNDARIES.map(idx =>
      (x.getPixelForValue(idx) + x.getPixelForValue(idx + 1)) / 2
    );

    ctx.save();
    ctx.strokeStyle = "rgba(245, 165, 36, 0.55)";
    ctx.lineWidth = 1.5;
    ctx.setLineDash([6, 4]);
    linePositions.forEach(pos => {{
      ctx.beginPath();
      ctx.moveTo(pos, top);
      ctx.lineTo(pos, bottom);
      ctx.stroke();
    }});

    ctx.setLineDash([]);
    ctx.fillStyle = "rgba(245, 165, 36, 0.75)";
    ctx.font = "bold 11px Segoe UI, sans-serif";
    ctx.textAlign = "center";

    const segments = [left, ...linePositions, right];
    const labels = ["2024", "2025", "2026"];
    labels.forEach((label, i) => {{
      const center = (segments[i] + segments[i + 1]) / 2;
      ctx.fillText(label, center, top + 16);
    }});
    ctx.restore();
  }},
}};

const ctx = document.getElementById("chart").getContext("2d");
const chart = new Chart(ctx, {{
  type: "line",
  data: {{ labels: PHASES, datasets }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    interaction: {{ mode: "nearest", intersect: true }},
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{
        backgroundColor: "rgba(255,255,255,0.97)",
        borderColor: "rgba(0,0,0,0.15)",
        borderWidth: 1,
        titleColor: "#111111",
        bodyColor: "#333333",
        padding: 14,
        callbacks: {{
          label(ctx) {{
            const v = ctx.parsed.y;
            if (v === null) return null;
            return ` ${{ctx.dataset.label}}: ${{v.toFixed(1)}}`;
          }},
        }},
      }},
    }},
    scales: {{
      x: {{
        grid: {{ color: "rgba(0,0,0,0.07)" }},
        ticks: {{ color: "#555555", font: {{ size: 11 }} }},
      }},
      y: {{
        grid: {{ color: "rgba(0,0,0,0.07)" }},
        ticks: {{
          color: "#555555",
          font: {{ size: 11 }},
          callback: v => v.toFixed(0),
        }},
        title: {{
          display: true,
          text: "Elo Rating",
          color: "#555555",
          font: {{ size: 12 }},
        }},
      }},
    }},
  }},
  plugins: [boundaryPlugin],
}});

const regionFiltersEl = document.getElementById("regionFilters");
const legendEl = document.getElementById("legend");

function syncVisibility() {{
  TEAMS.forEach((team, idx) => {{
    const showByRegion = activeRegions.has(team.region);
    chart.data.datasets[idx].hidden = !showByRegion || manualHidden[idx];
  }});
  chart.update();
}}

function renderRegionFilters() {{
  regionFiltersEl.innerHTML = "";

  const allBtn = document.createElement("button");
  allBtn.className = "region-filter" + (activeRegions.size === REGIONS.length ? " active" : "");
  allBtn.textContent = "All";
  allBtn.addEventListener("click", () => {{
    activeRegions = new Set(REGIONS);
    renderRegionFilters();
    renderLegend();
    syncVisibility();
  }});
  regionFiltersEl.appendChild(allBtn);

  REGIONS.forEach(region => {{
    const btn = document.createElement("button");
    btn.className = "region-filter" + (activeRegions.has(region) ? " active" : "");
    btn.textContent = region;
    btn.addEventListener("click", () => {{
      if (activeRegions.has(region)) {{
        activeRegions.delete(region);
      }} else {{
        activeRegions.add(region);
      }}
      renderRegionFilters();
      renderLegend();
      syncVisibility();
    }});
    regionFiltersEl.appendChild(btn);
  }});
}}

function renderLegend() {{
  legendEl.innerHTML = "";
  TEAMS.forEach((t, i) => {{
    if (!activeRegions.has(t.region)) return;
    const item = document.createElement("div");
    item.className = "legend-item" + (manualHidden[i] ? " hidden" : "");
    item.dataset.idx = i;
    item.innerHTML = `
      <span class="dot" style="background:${{t.color}}"></span>
      <span class="legend-name">${{t.name}}</span>
      <span class="legend-elo">${{t.finalElo !== null ? t.finalElo.toFixed(1) : "—"}}</span>
    `;
    item.addEventListener("click", () => {{
      manualHidden[i] = !manualHidden[i];
      item.classList.toggle("hidden", manualHidden[i]);
      syncVisibility();
    }});
    legendEl.appendChild(item);
  }});
}}

renderRegionFilters();
renderLegend();
syncVisibility();
</script>
</body>
</html>"""


def main() -> None:
    print("=== OWCS Combined Elo Chart (2024-2026, 2026 active teams only) ===")
    print(f"  Team history : {TEAM_HISTORY_XLSX.name}")
    print(f"  History files: {HISTORY_2024.name}, {HISTORY_2025.name}, {HISTORY_2026.name}")
    print(f"  Phases       : {len(PHASE_DEFS)}")

    teams = load_data()
    print(f"  Teams loaded : {len(teams)}")
    print("  Top 15 latest Elo:")
    for team in teams[:15]:
        print(f"    {team['name']:<28} {team['finalElo']:.1f}")

    html = build_html(teams)
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"\n  Generated: {OUTPUT_HTML}")


if __name__ == "__main__":
    main()
